# --- DISCLAIMER: I made this utilizing GPT-5, Gemini.
import os
import sys
import subprocess
import datetime
import pandas as pd

# Ensure required packages are installed
def install_if_missing(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

for pkg in ["mutagen", "pandas", "openpyxl"]:
    install_if_missing(pkg)

from mutagen.easyid3 import EasyID3
from mutagen.mp3 import MP3
from openpyxl import load_workbook
from openpyxl.styles import Font

# Hardcoded folders to avoid
EXCLUDE_FOLDERS = {'sound effects', 'Playlists', '2. Audiobooks'}

# Hardcoded column widths (adjust here as needed)
# Keys = column names, Values = desired Excel width
COLUMN_WIDTHS = {
    "Title": 50,
    "Contributing Artists": 30,
    "Album Artist": 25,
    "Album": 25,
    "Year": 8,
    "#": 6,
    "Genre": 15,
    "Length": 8,
    "MB": 8,
    "kbps": 8,
    "Full Path": 100,
    "Length/Size Ratio": 20
}

# Prompt for directory to scan
scan_path = input("Enter the directory path to scan for MP3 files: ").strip()

if not os.path.isdir(scan_path):
    print("Invalid path. Exiting.")
    sys.exit(1)

# Function to format seconds into mm:ss (supports >60 mins)
def format_duration(seconds):
    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    return f"{minutes}:{secs:02d}"

# List to store MP3 info
mp3_data = []

# Walk through directories
for root, dirs, files in os.walk(scan_path):
    dirs[:] = [d for d in dirs if d not in EXCLUDE_FOLDERS]
    
    for file in files:
        if file.lower().endswith(".mp3"):
            file_path = os.path.join(root, file)
            try:
                audio = MP3(file_path)
                tags = EasyID3(file_path)

                title = tags.get("title", [""])[0]
                artist = tags.get("artist", [""])[0]
                album_artist = tags.get("albumartist", [""])[0]
                album = tags.get("album", [""])[0]
                year = tags.get("date", [""])[0]
                track_num = tags.get("tracknumber", [""])[0]
                genre = tags.get("genre", [""])[0]
                length_str = format_duration(audio.info.length)
                size_mb = round(os.path.getsize(file_path) / (1024 * 1024), 2)  # MB
                bitrate = round(audio.info.bitrate / 1000)  # kbps
                ratio = audio.info.length/size_mb

                mp3_data.append({
                    "Title": title,
                    "Contributing Artists": artist,
                    "Album Artist": album_artist,
                    "Album": album,
                    "Year": year,
                    "#": track_num,
                    "Genre": genre,
                    "Length": length_str,
                    "MB": size_mb,
                    "kbps": bitrate,
                    "Full Path": file_path,  # store raw path, style later
                    "Length/Size Ratio": ratio
                })
            except Exception as e:
                print(f"Error reading {file_path}: {e}")

# Create DataFrame
df = pd.DataFrame(mp3_data)

# Timestamp for filename
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f"CompleteList{timestamp}.xlsx"

# Save to Excel
df.to_excel(output_filename, index=False, engine='openpyxl')

# Load workbook to adjust styles
wb = load_workbook(output_filename)
ws = wb.active

# Apply hyperlink style (blue + underline) and convert to clickable
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=ws.max_column)
    path = cell.value
    cell.value = path
    cell.hyperlink = path
    cell.font = Font(color="0000FF", underline="single")

# Apply fixed column widths from COLUMN_WIDTHS
for col in ws.iter_cols(min_row=1, max_row=1):
    header = col[0].value
    if header in COLUMN_WIDTHS:
        ws.column_dimensions[col[0].column_letter].width = COLUMN_WIDTHS[header]

wb.save(output_filename)

print(f"\nScan complete! Data saved to {output_filename}")
